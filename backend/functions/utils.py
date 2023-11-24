import csv
import io

from openpyxl import load_workbook
import pandas as pd


def get_label_hierarchy(file_content: bytes, file_type: str) -> list:
    indentation_list = []
    if file_type == "csv":
        def csv_indentation(cell_value):
            return len(cell_value) - len(cell_value.lstrip())

        with io.StringIO(file_content.decode('utf-8')) as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                # Assuming the first column may have indentation
                indentation_list.append(int(csv_indentation(row[0])))

    elif file_type == "xlsx":
        # Load an Excel file
        wb = load_workbook(filename=io.BytesIO(file_content))
        sheet = wb['Data']
        # Iterate through the rows and get indentation levels
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
            for cell in row:
                indentation_list.append(int(cell.alignment.indent))

    return indentation_list


def get_last_header_idx(df: pd.DataFrame) -> int:
    for row in df.itertuples(index=True):
        if 'Label' in row:
            return row[0]

    return 0


def check_for_multiheader(df: pd.DataFrame) -> bool:
    count = 0
    for col in df.columns:
        if "Unnamed" in col:
            count += 1

    if len(df.columns) - count <= 1:
        return True

    return False


def squash_column_header(df: pd.DataFrame, label_idx: int) -> list:
    squash_columns = []

    for i in range(label_idx + 1):
        row = list(df.iloc[i])
        temp_str = ""
        temp = []
        for value in row:
            if temp_str != "" and pd.isna(value):
                temp.append(temp_str)
                temp_str = ""
                continue
            elif not pd.isna(value):
                temp_str += str(value)

            temp.append(str(value))

        squash_columns.append(temp)

    header_list = []
    for values in zip(*squash_columns):
        header_column = ' '.join(element for element in values if element != 'nan')
        header_list.append(header_column)

    return header_list


def sanitize_columns(df: pd.DataFrame, indentation: list):
    if check_for_multiheader(df):
        label_idx = get_last_header_idx(df)
        header_list = squash_column_header(df, label_idx)
        df.columns = header_list
        df["Indentation"] = indentation
        df.drop(df.index[:label_idx + 1], inplace=True)
        df.reset_index(drop=True, inplace=True)
    else:
        df["Indentation"] = indentation

def add_label_grouping(df: pd.DataFrame):
    recent_labels = {}

    for idx, row in df.iterrows():
        current_indentation = row['Indentation']
        label = row['Label'].strip()

        # Update the dictionary with the current label for the current indentation level
        recent_labels[current_indentation] = label

        # Concatenate with the immediate lower level label if indentation is greater than 1
        if current_indentation > 1:
            parent_indentation = current_indentation - 1
            parent_label = recent_labels.get(parent_indentation, '')
            new_label = parent_label + " " + label
            df.at[idx, 'Label'] = new_label

